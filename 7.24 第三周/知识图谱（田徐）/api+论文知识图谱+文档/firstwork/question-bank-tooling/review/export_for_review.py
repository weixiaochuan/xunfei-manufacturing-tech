"""
人工审核导出/回写脚本（升级版：解决"每次导出都叫 pending_review.xlsx 会互相覆盖"的问题）。

Phase 1 不做审核界面，用 Excel 表格做人工审核。

—— 这次升级了什么 ——
1. 导出文件名自动带"章节/模型/时间戳"，每次导出都是新文件，绝不覆盖上一次。
   例：review/待审核_第四章_精度_ali_glm_20260710_1530.xlsx
   你原来手动 Copy-Item 改名的活儿，现在脚本自己干了。
2. 可以只导出某一章 / 某个模型出的题，而不是每次把全部待审核题一股脑倒出来。
3. 表格里多了两列 generation_model（哪个模型出的）、course_chapter（哪一章），
   审核时一眼能看出这题是 deepseek 还是星火出的、属于哪章，方便对比模型质量。
4. import_back 可以指定读哪个文件（因为现在文件不止一个了）。

—— 用法 ——
    # 导出"当前所有待审核/被驳回"的题（和以前行为一样，但文件名带时间戳不覆盖）
    python3 review/export_for_review.py export

    # 只导出第四章、ali_glm 出的题，并给文件起个好认的名字
    python3 review/export_for_review.py export --chapter 第四章_机械加工精度及其控制 --model ali_glm --name 精度混合

    # ... 人工在导出的那个 xlsx 里，把 review_status 列改成 已通过/已驳回 ...

    # 回写：默认回写 review/ 目录下"最新"的那个导出文件；也可 --file 指定
    python3 review/export_for_review.py import_back
    python3 review/export_for_review.py import_back --file "review/待审核_第四章_精度_ali_glm_20260710_1530.xlsx"
"""
import argparse
import glob
import os
import sqlite3
from datetime import datetime

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
import sys as _qb_sys
_qb_sys.path.insert(0, BASE_DIR)
from qb_runtime import connect_database, database_path
DB_PATH = database_path()
REVIEW_DIR = os.path.join(BASE_DIR, "review")

# 导出列：比原来多了 generation_model / course_chapter，审核时能区分模型和章节
COLUMNS = [
    "question_id", "course_chapter", "generation_model", "question_type",
    "source_node_id", "stem", "options_json", "answer", "explanation",
    "bloom_level", "calc_verify_status", "calc_verify_detail",
    "review_status", "review_comment",
]


def _sanitize(text):
    """把不能做文件名的字符去掉。"""
    bad = '\\/:*?"<>| '
    return "".join(c for c in str(text) if c not in bad)


def export(conn, chapter=None, model=None, name=None):
    select_cols = [c for c in COLUMNS if c != "review_comment"]
    sql = (
        f"SELECT {', '.join(select_cols)} FROM questions "
        f"WHERE review_status IN ('待审核','已驳回')"
    )
    params = []
    if chapter:
        sql += " AND course_chapter=?"; params.append(chapter)
    if model:
        # 允许只写 ali_glm，不用写全 ali_glm/glm-5.2
        sql += " AND generation_model LIKE ?"; params.append(f"{model}%")
    sql += " ORDER BY created_at"
    rows = conn.execute(sql, params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "待审核题目"
    ws.append(COLUMNS)
    for row in rows:
        ws.append(list(row) + [""])  # review_comment 留空给人工填

    # review_status 列加下拉校验，减少人工填错
    status_col_idx = COLUMNS.index("review_status") + 1
    dv = DataValidation(type="list", formula1='"待审核,已通过,已驳回"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{ws.cell(row=2, column=status_col_idx).coordinate}:"
           f"{ws.cell(row=ws.max_row, column=status_col_idx).coordinate}")

    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 60)

    # ---- 关键：文件名自动带标签+时间戳，永不覆盖 ----
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    parts = ["待审核"]
    if name:
        parts.append(_sanitize(name))
    if chapter:
        parts.append(_sanitize(chapter))
    if model:
        parts.append(_sanitize(model))
    parts.append(stamp)
    out_path = os.path.join(REVIEW_DIR, "_".join(parts) + ".xlsx")

    os.makedirs(REVIEW_DIR, exist_ok=True)
    wb.save(out_path)
    print(f"已导出 {len(rows)} 道题 -> {out_path}")
    print("（这是一个新文件，不会覆盖你之前导出的任何表格。审核完用 import_back 回写。）")


def _latest_review_file():
    files = glob.glob(os.path.join(REVIEW_DIR, "*.xlsx"))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def import_back(conn, file_path=None):
    if not file_path:
        file_path = _latest_review_file()
        if not file_path:
            print(f"{REVIEW_DIR} 里没有找到任何审核表格，请先运行 export")
            return
        print(f"未指定 --file，自动回写最新的一个：{os.path.basename(file_path)}")
    if not os.path.exists(file_path):
        print(f"未找到文件 {file_path}")
        return

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        qid = row[idx["question_id"]]
        status = row[idx["review_status"]]
        if not qid or status not in ("已通过", "已驳回"):
            continue
        conn.execute("UPDATE questions SET review_status=? WHERE question_id=?", (status, qid))
        updated += 1
    conn.commit()
    print(f"已回写 {updated} 道题目的审核结果（来源文件：{os.path.basename(file_path)}）")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("action", choices=["export", "import_back"])
    parser.add_argument("--chapter", help="只导出某一章，如 第四章_机械加工精度及其控制")
    parser.add_argument("--model", help="只导出某个模型出的题，如 ali_glm / deepseek / xinghuo")
    parser.add_argument("--name", help="给这次导出的文件起个好认的名字，如 精度混合")
    parser.add_argument("--file", help="import_back 时指定回写哪个文件；不填则回写最新的")
    args = parser.parse_args()

    conn = connect_database()
    if args.action == "export":
        export(conn, chapter=args.chapter, model=args.model, name=args.name)
    else:
        import_back(conn, file_path=args.file)
    conn.close()


if __name__ == "__main__":
    main()
