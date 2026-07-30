"""
自动化审核脚本（Auto-review）—— 让 AI 按【固定审核标准】先审一遍，
只把"有问题"的题捞出来给人看，通过的直接入库。这样批量出题才可行，
人只需复核那 15%~20% 的问题题，工程量大幅下降。

审核标准来自你定的《审核标准》，固定成机器可执行的规则，保证每题一把尺子：

【审核结果·四级 + 淘汰】（用于统计通过率）
    ✅ 通过
    🟢 修改后通过（轻微）
    🟡 修改后通过（中度）
    🟠 修改后通过（大改）
    ❌ 不合格（建议重写）
  通过率 =（通过 + 三档修改后通过）÷ 总题数

【问题类型·统一标签（可多选，用分号隔开）】
    知识点错误 / 题干歧义 / 条件不足 / 数据错误 / 公式错误 / 计算错误 /
    单位错误 / 逻辑错误 / 选项设计不合理 / 答案错误 / 解析错误 /
    难度不合理 / AI幻觉 / 教材依据不足 / 其他

【审核五步法（写进给模型的规则里）】
    1 判断：通过 / 修改 / 重写
    2 定位问题（打统一标签）
    3 给修改建议（为什么改）
    4 若需要则直接重写题目（保持知识点一致、难度一致、符合教材）
    5 再次验算，确保改后没有新问题

它做两件事：
  A) 对数据库里 review_status='待审核' 的题批量自动审；
  B) 把结果写成和人工审核一样结构的 Excel（原列 + 审核结果/问题类型/修改建议/
     修改后题目/修改后答案），既能直接 import_reviewed.py 回库，也能当论文里的
     "人工审核日志"。同时把"✅通过 / 🟢轻微"直接标为已通过写回库，只有中度及以上
     留成 Excel 交人复核。

强烈建议用能力强的模型做审核员（deepseek 即可），审核比出题更需要判断力。

用法：
    # 审当前所有待审核题，导出问题题到 Excel，通过的直接入库
    python3 review/auto_review.py --provider deepseek

    # 只审某一章
    python3 review/auto_review.py --provider deepseek --chapter 第三章_机床夹具设计

    # 试跑（只审5道，看看效果）
    python3 review/auto_review.py --provider deepseek --limit 5
"""
import argparse
import json
import os
import sqlite3
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
import sys as _qb_sys
_qb_sys.path.insert(0, BASE_DIR)
from qb_runtime import connect_database, database_path
DB_PATH = database_path()
REVIEW_DIR = os.path.join(BASE_DIR, "review")

GRADES = ["✅ 通过", "🟢 修改后通过（轻微）", "🟡 修改后通过（中度）",
          "🟠 修改后通过（大改）", "❌ 不合格（建议重写）"]
PROBLEM_TAGS = ["知识点错误", "题干歧义", "条件不足", "数据错误", "公式错误", "计算错误",
                "单位错误", "逻辑错误", "选项设计不合理", "答案错误", "解析错误",
                "难度不合理", "AI幻觉", "教材依据不足", "其他"]

REVIEW_SYSTEM = """你是一位机械制造工程学科的资深命题审核专家，正在按统一标准审核 AI 出的题。你的判断要严格、就事论事，宁可严一点也不放水——因为通过的题会直接给学生用。

【审核五步法】
1 判断该题应：通过 / 修改后通过 / 不合格重写。
2 定位问题，只能从下面这套【统一标签】里选（可多选，用中文分号；隔开）：
   知识点错误；题干歧义；条件不足；数据错误；公式错误；计算错误；单位错误；逻辑错误；选项设计不合理；答案错误；解析错误；难度不合理；AI幻觉；教材依据不足；其他
3 给出修改建议（说清为什么改）。
4 若是中度/大改/重写，直接给出修改后的题目与修改后的答案（保持知识点一致、难度一致、符合教材原文）。
5 复核：确保你给的修改本身没有新问题（尤其计算题要把算式重新验算一遍）。

【审核结果·必须从这五个里选一个原样输出】
✅ 通过
🟢 修改后通过（轻微）
🟡 修改后通过（中度）
🟠 修改后通过（大改）
❌ 不合格（建议重写）

【判档尺度】
- ✅ 通过：题干清楚、条件充分、答案与解析正确自洽、选项设计合理（单选恰好1个正确项且干扰项都对应真实误区）、计算题步骤可验算且数值正确。
- 🟢 轻微：仅个别措辞/标点/解析表述小瑕疵，不影响答题。无需给修改后题目。
- 🟡 中度：有明确但局部的错误（如某处数据/单位/解析小错），改一处即可。给出修改后题目/答案。
- 🟠 大改：题干或选项结构性问题、需要较大改动才能用。给出修改后题目/答案。
- ❌ 重写：编造条件/公式（AI幻觉）、答案错误且难救、内容空洞无考察价值、无正确选项或多正确选项且逻辑崩坏。

【只输出一个JSON对象，不要用Markdown代码块包裹】
{
  "grade": "上面五个之一，原样",
  "problem_tags": "命中的问题标签，用；分隔；若通过则填 无",
  "suggestion": "修改建议；若通过则填 无",
  "revised_stem": "修改后题目；仅中度/大改/重写填写，否则留空字符串",
  "revised_answer": "修改后答案（含步骤/解析）；仅中度/大改/重写填写，否则留空字符串"
}"""

REVIEW_USER = """请审核下面这道题。

【题型】{qtype}
【所属章节】{chapter}
【来源知识点ID】{node}
【题干】
{stem}

【选项 或 计算步骤】
{options}

【标注答案】
{answer}

【解析】
{explanation}

【计算题自动验算状态】{calc_status}　{calc_detail}

【该知识点原文】（判断"是否符合教材""是否幻觉"的依据）
{content}

请按审核五步法给出JSON。"""


def _content(conn, node):
    r = conn.execute("SELECT knowledge_title, content, formulas FROM knowledge_points WHERE knowledge_id=?",
                     (node,)).fetchone()
    if not r:
        return ""
    c = r[1] or ""
    if r[2]:
        c += f"\n【相关公式】{r[2]}"
    return c


def review_one(client, conn, q):
    from pipeline.generate_questions import parse_llm_json
    opts = q["options_json"] or ""
    user = REVIEW_USER.format(
        qtype=q["question_type"], chapter=q["course_chapter"], node=q["source_node_id"],
        stem=q["stem"], options=str(opts)[:1500], answer=q["answer"],
        explanation=q["explanation"] or "", calc_status=q["calc_verify_status"] or "",
        calc_detail=q["calc_verify_detail"] or "", content=_content(conn, q["source_node_id"])[:2500])
    raw = client.chat(REVIEW_SYSTEM, user, temperature=0.2)
    data = parse_llm_json(raw)
    grade = data.get("grade", "").strip()
    if grade not in GRADES:
        # be forgiving: map by keyword
        for g in GRADES:
            if g[-2:] in grade or g[0] in grade:
                grade = g
                break
        else:
            grade = "🟡 修改后通过（中度）"
    return {
        "grade": grade,
        "problem_tags": data.get("problem_tags", ""),
        "suggestion": data.get("suggestion", ""),
        "revised_stem": data.get("revised_stem", ""),
        "revised_answer": data.get("revised_answer", ""),
    }


EXPORT_COLS = ["question_id", "course_chapter", "generation_model", "question_type",
               "source_node_id", "stem", "options_json", "answer", "explanation",
               "bloom_level", "calc_verify_status", "calc_verify_detail",
               "review_status", "审核结果", "问题类型", "修改建议", "修改后题目", "修改后答案"]

# 直接放行的档位（写回库为已通过）；其余留人复核
AUTO_PASS = {"✅ 通过", "🟢 修改后通过（轻微）"}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--provider", required=True, help="做审核员的模型，如 deepseek")
    ap.add_argument("--chapter", default=None)
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--dry_run", action="store_true", help="只审不写库（仍导出Excel）")
    a = ap.parse_args()

    from llm.client import get_client
    client = get_client("concept", provider_name=a.provider)

    conn = connect_database()
    conn.row_factory = sqlite3.Row
    sql = "SELECT * FROM questions WHERE review_status='待审核'"
    params = []
    if a.chapter:
        sql += " AND course_chapter=?"; params.append(a.chapter)
    sql += " ORDER BY created_at"
    if a.limit:
        sql += " LIMIT ?"; params.append(a.limit)
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    if not rows:
        print("没有待审核的题（review_status='待审核'）。")
        return

    print(f"开始自动审核 {len(rows)} 道题，审核员模型：{a.provider}\n")
    results, stats = [], {}
    for i, q in enumerate(rows, 1):
        try:
            r = review_one(client, conn, q)
        except Exception as e:
            r = {"grade": "🟡 修改后通过（中度）", "problem_tags": "其他",
                 "suggestion": f"自动审核出错，转人工：{e}", "revised_stem": "", "revised_answer": ""}
        results.append((q, r))
        stats[r["grade"]] = stats.get(r["grade"], 0) + 1
        print(f"[{i}/{len(rows)}] {q['question_id']} {q['question_type']} -> {r['grade']}"
              + (f"  ({r['problem_tags']})" if r['problem_tags'] and r['problem_tags'] != '无' else ""))
        # 写回库：AUTO_PASS 直接标已通过；其余保持待审核，交人看Excel
        if not a.dry_run and r["grade"] in AUTO_PASS:
            conn.execute("UPDATE questions SET review_status='已通过', common_error_tags=? WHERE question_id=?",
                         (f"auto:{r['grade']}", q["question_id"]))
    if not a.dry_run:
        conn.commit()

    # 导出"需人复核"的题到 Excel（中度及以上），文件名带时间戳不覆盖
    need_human = [(q, r) for q, r in results if r["grade"] not in AUTO_PASS]
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    tag = ("_" + a.chapter) if a.chapter else ""
    out = os.path.join(REVIEW_DIR, f"auto_review_需复核{tag}_{stamp}.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "自动审核-需复核"
    ws.append(EXPORT_COLS)
    for q, r in need_human:
        ws.append([q["question_id"], q["course_chapter"], q["generation_model"], q["question_type"],
                   q["source_node_id"], q["stem"], q["options_json"], q["answer"], q["explanation"],
                   q["bloom_level"], q["calc_verify_status"], q["calc_verify_detail"],
                   q["review_status"], r["grade"], r["problem_tags"], r["suggestion"],
                   r["revised_stem"], r["revised_answer"]])
    # 下拉校验
    gi = EXPORT_COLS.index("审核结果") + 1
    dv = DataValidation(type="list", formula1='"' + ",".join(GRADES) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    if ws.max_row >= 2:
        dv.add(f"{ws.cell(row=2, column=gi).coordinate}:{ws.cell(row=ws.max_row, column=gi).coordinate}")
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 60)
    os.makedirs(REVIEW_DIR, exist_ok=True)
    wb.save(out)
    conn.close()

    total = len(results)
    passed = sum(v for k, v in stats.items() if k in AUTO_PASS)
    effective = sum(v for k, v in stats.items() if k != "❌ 不合格（建议重写）")
    print("\n===== 自动审核汇总 =====")
    for g in GRADES:
        print(f"  {g}: {stats.get(g, 0)}")
    print(f"  ------")
    print(f"  自动放行(已入库): {passed} 道")
    print(f"  需人工复核(导出Excel): {len(need_human)} 道 -> {out}")
    print(f"  原始通过率: {passed/total*100:.1f}%  有效通过率(含可改): {effective/total*100:.1f}%")
    print("\n下一步：打开那个 Excel 复核'需复核'的题，改好后用")
    print(f"  python3 review/import_reviewed.py --file \"{out}\"  回写入库。")


if __name__ == "__main__":
    main()
